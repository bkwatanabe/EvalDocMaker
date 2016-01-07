from openpyxl import load_workbook
from datetime import datetime
import json

wb = load_workbook("sample knee patients.xlsx")
ws = wb.active
rows = ws.rows


def getPatientInfo():

    patients = []

    for x in range(1,ws.max_row):
        if(rows[x][1].value is not None):
            patient = {"id":x}
            for y in range(0, ws.max_column):
                value = rows[x][y].value
                if(value == '-'):
                    value = None
                if(isinstance(value,datetime)):
                    format = '%m-%d-%Y'
                    value = str(value.date().strftime(format))
                patient[rows[0][y].value] = value
            patients.append(patient)
    return patients

def getJsonOfPatientInfo():
    return json.dumps(getPatientInfo(), indent=4, separators=(',', ': '))

print getPatientInfo()
print getJsonOfPatientInfo()